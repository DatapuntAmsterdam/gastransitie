import os
import logging

import pandas
from openpyxl import load_workbook
from sqlalchemy import create_engine
from django.db import connections
from django.contrib.gis.geos import MultiPolygon
from django.contrib.gis.geos import GEOSGeometry
from datasets.models import bag
from datasets.models import alliander

from datasets.imports.util import zipped_shp2psql
from datasets.imports.util import get_sqlalchemy_db_url
from datasets.imports.util import get_ogr2ogr_pgstr
from datasets.imports.util import SRS_TO_STORE


log = logging.getLogger(__name__)


def _load_stoplicht_alliander(filename):
    """
    Load the initial Alliander stoplicht data set.
    """
    # Excel wrangling:
    sheetname = "Buurten Stoplicht Liander"
    worksheet = load_workbook(filename=filename)[sheetname]
    columns = [x.value for x in worksheet[1]]
    data = ([x.value for x in row] for row in worksheet.iter_rows(min_row=2))

    df = pandas.DataFrame(data, columns=columns)

    # Write our data to database
    engine = create_engine(get_sqlalchemy_db_url())
    df.to_sql('gas_alliander_stoplicht_raw', engine, if_exists='replace')


def load_xslx_verbruik_kv(datadir):
    """
    Verbruik P6 data. (klein)
    """
    # LianderKV01012018.xlsx
    xls_path = os.path.join(datadir, 'alliander', 'LianderKV01012018.xlsx')
    log.debug('Loading worksheet')
    sheetname = "Export Worksheet"
    wb = load_workbook(filename=xls_path, read_only=True)
    worksheet = wb[sheetname]
    worksheet.max_row = worksheet.max_column = None
    columns = [x.value for x in worksheet[1]]
    # replace invalid sql chars
    columns = [x.replace('%', '') for x in columns]
    columns = [x.replace('(', ' ') for x in columns]
    columns = [x.replace(')', ' ') for x in columns]

    log.debug('Columns data %s', columns)
    log.debug('Loading data')

    data = []

    for ri, row in enumerate(worksheet.iter_rows(min_row=2)):
        if ri % 5000 == 0:
            log.debug('rows..%d', ri)

        # Filter out non Amsterdam values
        if row[6].value != 'AMSTERDAM':
            continue

        data.append([x.value for x in row])

    log.debug('Create pandas df data')
    df = pandas.DataFrame(data, columns=columns)
    engine = create_engine(get_sqlalchemy_db_url())
    log.debug('dump to sql')
    df.to_sql('alliander_kv', engine, if_exists='replace')


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]


def create_p6_buurt():
    sql = f"""
    SELECT DISTINCT postcode, b.id as id, b.vollcode as code, b.naam as naam
    FROM bag_nummeraanduiding n, bag_verblijfsobject v, bag_buurt b
    WHERE n.verblijfsobject_id = v.id
    AND v.buurt_id = b.id
    AND postcode is not null
    AND postcode != ''
    ORDER BY naam
    """

    with connections['bag'].cursor() as cursor:
        cursor.execute(sql)
        data = dictfetchall(cursor)

    postcode_buurt = {}

    for item in data:
        postcode_buurt[item['postcode']] = {
            'naam': item['naam'],
            'id': item['id'],
            'vollcode': item['code'],
        }

    return postcode_buurt


def create_p6_panden():
    """
    For each P6 collect geometrie and id's of involved PANDEN.
    """
    sql = f"""
    SELECT
        p.id, p.geometrie, count(*),
        array_agg(DISTINCT n.postcode) as postcodes,
        array_agg(DISTINCT v.buurt_id) as buurtids
    FROM bag_nummeraanduiding n, bag_pand p,
         bag_verblijfsobject v,
         bag_verblijfsobjectpandrelatie vp,
         bag_buurt b
    WHERE n.verblijfsobject_id = v.id
    AND vp.pand_id = p.id
    AND vp.verblijfsobject_id = v.id
    AND v.buurt_id = b.id
    GROUP BY (p.id)
    """

    log.debug('Extract Panden with postcodes..')
    with connections['bag'].cursor() as cursor:
        cursor.execute(sql)
        # ~146.000
        data = dictfetchall(cursor)

    p6_panden = {}

    for pand in data:
        for p6 in pand['postcodes']:
            if p6 not in p6_panden:
                p6_panden[p6] = [pand]
            else:
                p6_panden[p6].append(pand)

    return p6_panden


def link_panden_raport(p6_rapport, panden):
    """link pand data <-> p6 rapport a
    """
    for pand in panden:
        p6_rapport['postcodes'].update(pand['postcodes'])
        p6_rapport['buurt_ids'].update(pand['buurtids'])
        # p6_rapport['panden_geo'].append(pand)
        p6_rapport['panden'].append(pand)

        # make sure every pand can find out
        # in which rapports it is invoved with.
        # in most cases just 1. BUT some big buildings have many
        # P6 codes / rapports within them
        rapporten = pand.get('rapporten', [])
        rapporten.append(p6_rapport)
        pand['rapporten'] = rapporten


def make_empty_p6_rapport():
    """Template to increase for each pand in p6 range
    """
    return {
        'postcodes': set(),
        'buurt_ids': set(),
        'rapport_count': 0,
        'panden': [],
        'gas': {
            'aansluitingen': 0,
            'm3': 0,
            'leveringsrichting': 0
        },
        'elk': {
            'aansluitingen': 0,
            'Kwh': 0,
            'leveringsrichting': 0
        }
    }


def make_empty_rapports():
    """Create empty energy rapports for each buurt
    """
    buurten = {}

    # make empty rapports.
    for b in bag.BagBuurt.objects.using('bag').all().order_by('naam'):
        buurt_rapport = {
            'naam': b.naam,
            'id': b.id,
            'vollcode': b.vollcode,
            'gas': {
                'aansluitingen': 0,
                'm3': 0,
                'leveringsrichting': 0
            },
            'elk': {
                'aansluitingen': 0,
                'Kwh': 0,
                'leveringsrichting': 0
            }
        }
        buurten[b.id] = buurt_rapport

    return buurten


def generate_postcodes(postcode_van, postcode_tot):
    """
    """
    if postcode_van == postcode_tot:
        return [postcode_van]
    # get middle postcodes.
    sql = """
    SELECT postcode from bag_nummeraanduiding n
    WHERE n.postcode >= '{postcode_van}'
    AND n.postcode <= '{postcode_tot}'
    """

    with connections['bag'].cursor() as cursor:
        cursor.execute(sql)
        data = dictfetchall(cursor)

    return [d['postcode'] for d in data]


def update_rapport(rapport, p6data):
    """Update rapport with energy usage information
    """
    r = rapport
    if p6data.productsoort == 'GAS':
        r['gas']['aansluitingen'] += p6data.aantal_aansluitingen
        r['gas']['m3'] += p6data.sjv
        r['gas']['leveringsrichting'] += p6data.leveringsrichting
    else:
        r['elk']['aansluitingen'] += p6data.aantal_aansluitingen
        r['elk']['Kwh'] += p6data.sjv
        r['elk']['leveringsrichting'] += p6data.leveringsrichting


def add_panden_to_kv(p6_panden):
    """
    Add pand information to all P6 alliander usage information
    """
    log.info('Add pand information to alliander kv p6 records')
    p6_rapports = {}

    alliander.VerbruikPerPandenP6.objects.all().delete()

    matched = 1

    for p6data in alliander.AllianderKv.objects.all().order_by('postcode_van'):

        # find all postcodes van -> tot
        postcodes = generate_postcodes(
            p6data.postcode_van, p6data.postcode_tot)

        rapport = make_empty_p6_rapport()

        # add all panden information to this
        for p6 in postcodes:
            panden = p6_panden.get(p6)
            if panden:
                link_panden_raport(rapport, panden)
                update_rapport(rapport, p6data)
                matched += len(panden)

        if matched % 1000 == 0:
            log.debug(matched)

        p6_rapports[p6data.postcode_van] = rapport

    logging.debug(matched)
    return p6_rapports


def _add_pand_to_collection(geomcollection, rapportcollection, rapport):

    rapportcollection.append(rapport)

    if rapport.get('visited'):
        return

    rapport['visited'] = True

    for pand in rapport['panden']:

        geomcollection.append(pand['geometrie'])

        if not pand['rapporten']:
            continue

        for sub_r in pand['rapporten']:
            _add_pand_to_collection(geomcollection, rapportcollection, sub_r)


def merge_rapport(m, r):
    m['rapport_count'] += 1
    m['gas']['aansluitingen'] += r['gas']['aansluitingen']
    m['gas']['m3'] += r['gas']['m3']
    m['gas']['leveringsrichting'] += r['gas']['leveringsrichting']
    m['elk']['aansluitingen'] += r['elk']['aansluitingen']
    m['elk']['Kwh'] += r['elk']['Kwh']
    m['elk']['leveringsrichting'] += r['elk']['leveringsrichting']


def sum_rapports(rapports):
    master = make_empty_p6_rapport()
    for r in rapports:
        merge_rapport(master, r)

    m = master

    m.pop('postcodes')
    m.pop('buurt_ids')

    return master


def create_usage_views():
    """Create sql views
    """

    sql_gas = """
CREATE OR REPLACE VIEW gas_verbruik AS
SELECT id, code, vollcode, buurt_id, geometrie,
CAST(data->'gas'->>'m3' as float) AS gasm3,
CAST(data->'gas'->>'aansluitingen' AS float) AS aansluitingen,
CAST(data->'gas'->>'leveringsrichting' as float) as leveringsrichting,
COALESCE(
    cast(data->'gas'->>'m3' as float) /
    nullif(cast(data->'gas'->>'aansluitingen' AS float), 0), 0) AS verbruik
FROM datasets_verbruikperpandenp6
    """

    sql_elk = """
CREATE OR REPLACE VIEW elk_verbruik as
SELECT id, code, vollcode, buurt_id, geometrie,
cast(data->'elk'->>'m3' as float) as gasm3,
cast(data->'elk'->>'aansluitingen' as float) as aansluitingen,
CAST(data->'elk'->>'leveringsrichting' as float) as leveringsrichting,
COALESCE(
    cast(data->'elk'->>'Kwh' as float) /
    nullif(cast(data->'elk'->>'aansluitingen' as float), 0), 0) as verbruik
FROM datasets_verbruikperpandenp6
    """

    with connections['default'].cursor() as cursor:
        cursor.execute(sql_gas)
        cursor.execute(sql_elk)


def create_panden_usage_map(p6_rapports):
    """
    For all 6p - pand relations create single
    usage information with multiple geometry objects
    """
    alliander.VerbruikPerPandenP6.objects.all().delete()
    log.info("Save usage information with Pand geometries")

    # makes this function idempotent
    for rapport in p6_rapports.values():
        rapport['visited'] = False

    for i, rapport in enumerate(p6_rapports.values()):
        if rapport.get('visited'):
            continue
        geomcollection = []
        collection_rapports = []

        _add_pand_to_collection(geomcollection, collection_rapports, rapport)

        geomcollection = [GEOSGeometry(p, srid=28992) for p in geomcollection]

        if not rapport['panden']:
            continue

        pand = rapport['panden'][0]
        buurt_id = pand['buurtids'][0]
        b = bag.BagBuurt.objects.using('bag').get(id=buurt_id)

        master_rapport = sum_rapports(collection_rapports)

        vp6 = alliander.VerbruikPerPandenP6(
            code=b.code,
            vollcode=b.vollcode,
            buurt_id=buurt_id,
            # postcode=
            geometrie=MultiPolygon(geomcollection),
            data=master_rapport
        )
        vp6.save()

        if i % 100 == 0:
            log.debug(i)

    count = alliander.VerbruikPerPandenP6.objects.all().count()
    log.info('P6 geo objects %d', count)


def verbruik_per_buurt(p6_buurt):
    """Stel gasverbruik per buurt vast.
    """
    alliander.VerbruikPerBuurt.objects.all().delete()
    b_rapports = make_empty_rapports()

    for p6data in alliander.AllianderKv.objects.all().order_by('postcode_van'):
        buurt = p6_buurt.get(p6data.postcode_van)
        if not buurt:
            log.debug('skipped %s', p6data.postcode_van)
            continue

        br = b_rapports[buurt['id']]

        if p6data.productsoort == 'GAS':
            br['gas']['aansluitingen'] += p6data.aantal_aansluitingen
            br['gas']['m3'] += p6data.sjv
            br['gas']['leveringsrichting'] += p6data.leveringsrichting
        else:
            br['elk']['aansluitingen'] += p6data.aantal_aansluitingen
            br['elk']['Kwh'] += p6data.sjv
            br['elk']['leveringsrichting'] += p6data.leveringsrichting

        avb = alliander.VerbruikPerBuurt(
            id=br['id'],
            vollcode=br['vollcode'],
            naam=br['naam'],
            data=br,
        )
        avb.save()


def import_alliander(datadir):
    """Import alliander data.

    -leidingen
    -verbruik
    -verbruik pand
    -verbruik buurt
    """
    load_xslx_verbruik_kv(datadir)
    p6_panden = create_p6_panden()
    p6_rapports = add_panden_to_kv(p6_panden)
    create_panden_usage_map(p6_rapports)
    create_usage_views()

    p6_buurt = create_p6_buurt()
    verbruik_per_buurt(p6_buurt)

    # stoplicht data
    # fn = '20170829 - Stoplicht Amsterdam DISTRIBUTIELEIDINGEN (deelbaar) v0.02.xlsx'  # noqa
    # _load_stoplicht_alliander(os.path.join(datadir, 'alliander', fn))

    # Alliander data for gas mains
    pg_str = get_ogr2ogr_pgstr()

    zipped_shp2psql(
        os.path.join(datadir, 'alliander', 'Groen_Amsterdam.zip'),
        'Groen_Amsterdam.shp',
        pg_str,
        'gas_alliander_gas_groen_raw',
        t_srs=SRS_TO_STORE,
        nlt='PROMOTE_TO_MULTI',
        s_srs='EPSG:28992'
    )

    zipped_shp2psql(
        os.path.join(datadir, 'alliander', 'Oranje_Amsterdam.zip'),
        'Oranje_Amsterdam.shp',
        pg_str,
        'gas_alliander_gas_oranje_raw',
        t_srs=SRS_TO_STORE,
        s_srs='EPSG:28992'
    )
